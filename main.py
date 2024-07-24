from tkinter import ttk, messagebox
import tkinter as tk
import json
import configparser
from datetime import time, datetime
import os
import sys
from redminelib import Redmine
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageTk
import pystray
from pystray import MenuItem as item


class App:
        
    # Update the path to be the current directory of the script
    os.chdir(os.path.dirname(os.path.abspath(sys.argv[0])))

    # Read configuration from config.ini
    config = configparser.ConfigParser()
    try:
        config.read('config.ini')
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read configuration file: {e}")
        sys.exit(1)

    # Load values from the configuration
    try:
        popup_interval = int(config['General']['popup_interval'])
        start_time = time(int(config['General']['start_time']), 0)
        end_time = time(int(config['General']['end_time']), 0)
        redmine_url = config['Redmine']['url']
        api_key = config['Redmine']['api_key']
        user_id = int(config['Redmine']['user_id'])
    except Exception as e:
        messagebox.showerror("Error", f"Invalid configuration: {e}")
        sys.exit(1)

    # Static variables
    TIME_STEP_REDMINE = popup_interval / 60 / 60
    SUB_PHASE_ID = 108
    ON_CALL_ID = 109
    CALL_IN_ID = 110
    FIVE_MINUTES = 60 * 5
    activities_dict = {
        '-': 27,
        'AD': 8,
        'CM': 9,
        'COMM': 10,
        'CU': 11,
        'DE': 12,
        'EX': 13,
        'HW': 14,
        'IT': 15,
        'MOI': 26,
        'OP': 16,
        'PM': 17,
        'PRO': 48,
        'QA': 18,
        'RM': 118,
        'SOP': 49,
        'SP': 19,
        'SR': 20,
        'SU': 21,
        'SW': 22,
        'VL': 23,
        'VR': 24,
        'WA': 25
    }

    def __init__(self, root):
        self.root = root

        # Define the interface elements and interactions
        self.root.title("PTT - Productivity tracking tool v24.06.24")
        self.root.attributes("-alpha", 0)

        # Remove window decorations
        root.overrideredirect(True)
        transparentcolor = 'grey'
        self.root.attributes("-transparentcolor", transparentcolor)

        # Load image
        bg_image_path = "background_clippy.png"
        bg_image = Image.open(bg_image_path)
        tk_image = ImageTk.PhotoImage(bg_image)
        
        # Keep a reference to the image to prevent it from being garbage collected
        root.tk_image = tk_image

        # Create a Canvas widget
        self.canvas = tk.Canvas(root, width=bg_image.width, height=bg_image.height, bg=transparentcolor, highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)
        
        # Display the image on the canvas
        self.canvas.create_image(0, 0, anchor=tk.NW, image=tk_image)
        x_pos = 205

        # Create the label
        self.label = ttk.Label(self.root, text="What are you doing my boi?", font=("Helvetica", 14), background="#ffffc9")
        self.label_window = self.canvas.create_window(x_pos, 15, anchor=tk.N, window=self.label)
        
        # Create the dropdown
        self.task_var = tk.StringVar()
        self.task_dropdown = ttk.Combobox(self.root, textvariable=self.task_var, font=("Helvetica", 9), width=50, state="readonly")
        self.task_dropdown_window = self.canvas.create_window(x_pos, 60, anchor=tk.N, window=self.task_dropdown)

        # Create the button frame
        self.button_frame = ttk.Frame(self.root, style="custom.TFrame", padding=5)
        self.button_frame_window = self.canvas.create_window(x_pos, 100, anchor=tk.N, window=self.button_frame)
        
        # Create the buttons
        self.confirm_button = ttk.Button(self.button_frame, text="Confirm", command=self.confirm, style="TButton")
        self.confirm_button.grid(row=0, column=0, padx=10)
        
        self.postpone_button = ttk.Button(self.button_frame, text="Postpone", command=self.postpone, style="TButton")
        self.postpone_button.grid(row=0, column=1, padx=10)

        # Styles
        # Create a custom style for the frame
        frame_style = ttk.Style()
        frame_style.configure("custom.TFrame", background="#ffffc9")

        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        # Main program logic
        self.task_list = []
        self.redmine_issues = {}
        
        self.timer = 0
        self.tray_icon = None
        self.show_tray_icon()
        self.on_tick()

        self.redmine = Redmine(self.redmine_url, key=self.api_key, requests={'verify': False})
        self.load_task_list()

        try:
            with open("daily_task_counts.json") as file:
                data = file.read()
                if data:
                    self.daily_task_counts = json.loads(data)
                else:
                    self.daily_task_counts = {}
        except FileNotFoundError:
            self.daily_task_counts = {}
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load daily task counts: {e}")
            self.daily_task_counts = {}

        self.check_and_show_popup()
    
            
    def save_task_list(self, task_list):
        try:
            with open('task_list.json', 'w') as file:
                json.dump(task_list, file, indent=4)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save task list: {e}")


    def add_new_task(self, task_list, task_name, phase_code, task_code):
        task_list['REDMINE_TASKS'][task_name] = {
            'phase_code': phase_code,
            'task_code': task_code
        }
        self.save_task_list(task_list)


    def check_task_list(self, task_name):
        # Check if its a custom task
        if task_name in self.task_list['CUSTOM_TASKS']:
            return self.task_list['CUSTOM_TASKS'][task_name]
        
        # Check if its a Redmine task
        if task_name in self.task_list['REDMINE_TASKS']:
            return self.task_list['REDMINE_TASKS'][task_name]
        else:
            def on_ok():
                try:
                    phase_code = phase_code_entry.get()
                    task_code = task_code_entry.get()
                    self.add_new_task(self.task_list, task_name, phase_code, task_code)
                    popup.destroy()
                    self.confirm()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to add new task: {e}")

            popup = tk.Toplevel()
            popup.title("Missing task info")
            tk.Label(popup, text="Please enter the phase and task codes for the new task").grid(row=0, columnspan=2, pady=(10, 10))
            
            tk.Label(popup, text="Phase Code").grid(row=1, column=0)
            tk.Label(popup, text="Task Code").grid(row=2, column=0)

            phase_code_entry = tk.Entry(popup)
            task_code_entry = tk.Entry(popup)

            phase_code_entry.grid(row=1, column=1)
            task_code_entry.grid(row=2, column=1)

            ok_button = tk.Button(popup, text="OK", command=on_ok)
            ok_button.grid(row=3, columnspan=2, pady=(10, 10))

            popup.mainloop()


    def load_redmine_issues(self):
        try:
            self.redmine_issues.clear()
            assigned_issues = self.redmine.issue.filter(assigned_to_id=self.user_id)

            for issue in assigned_issues:
                self.redmine_issues[issue.id] = issue.subject
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Redmine issues, ensure you have internet connection, guardian is online, and review the configuration values: {e}")


    def log_time_on_issue(self, issue_id):
        try:
            today = datetime.now().date()
            time_entries = self.redmine.time_entry.filter(issue_id=issue_id, spent_on=today, user_id=self.user_id)

            if time_entries:
                time_entry = time_entries[0]  # Assuming only one time entry per issue per day
                new_hours = time_entry.hours + self.TIME_STEP_REDMINE
                time_entry.hours = new_hours
                time_entry.save()
            else:
                redmine_issue = self.redmine_issues.get(issue_id)
                if redmine_issue:
                    sub_phase_value = self.task_list['REDMINE_TASKS'][redmine_issue]['task_code']
                    phase_code_value = self.task_list['REDMINE_TASKS'][redmine_issue]['phase_code']
                    activity_id = self.activities_dict[phase_code_value]
                    self.redmine.time_entry.create(
                        issue_id=issue_id,
                        spent_on=today,
                        hours=self.TIME_STEP_REDMINE,
                        activity_id=activity_id,
                        custom_fields=[
                            {'id': self.SUB_PHASE_ID, 'value': sub_phase_value},
                            {'id': self.ON_CALL_ID, 'value': ''},
                            {'id': self.CALL_IN_ID, 'value': ''}],
                        comments='Automated log time entry'
                    )
        except Exception as e:
            messagebox.showerror("Error", f"Failed to log time on issue: {e}")


    def confirm(self):
        try:
            now = datetime.now().strftime("%d-%m-%Y")
            task_name = self.task_var.get()
            task_details = self.check_task_list(task_name)
            phase_code = task_details["phase_code"]
            task_code = task_details["task_code"]
            self.timer = self.popup_interval

            # Create the structure for today
            if now not in self.daily_task_counts:
                self.daily_task_counts[now] = {
                    "REDMINE": {},
                    "NAVISION_TIMESHEET": {}
                }

            if task_name in self.daily_task_counts[now]["REDMINE"]:
                self.daily_task_counts[now]["REDMINE"][task_name] += self.TIME_STEP_REDMINE
            else:
                self.daily_task_counts[now]["REDMINE"][task_name] = self.TIME_STEP_REDMINE

            if phase_code in self.daily_task_counts[now]["NAVISION_TIMESHEET"]:
                if task_code in self.daily_task_counts[now]["NAVISION_TIMESHEET"][phase_code]:
                    self.daily_task_counts[now]["NAVISION_TIMESHEET"][phase_code][task_code] += self.TIME_STEP_REDMINE
                else:
                    self.daily_task_counts[now]["NAVISION_TIMESHEET"][phase_code][task_code] = self.TIME_STEP_REDMINE
            else:
                self.daily_task_counts[now]["NAVISION_TIMESHEET"][phase_code] = {task_code: self.TIME_STEP_REDMINE}

            with open("daily_task_counts.json", "w", encoding='utf-8') as file:
                json.dump(self.daily_task_counts, file, indent=4)

            # root.withdraw()
            self.hide()

            # If its a Redmine task, then log time
            issue_id = self.get_key_from_value(self.redmine_issues, task_name)
            if (issue_id):
                self.log_time_on_issue(issue_id)

            self.json_to_excel('daily_task_counts.json', 'output.xlsx')

            root.after(self.popup_interval * 1000, self.check_and_show_popup)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to confirm task: {e}")


    def get_key_from_value(self, in_dict, target_value):
        for key, value in in_dict.items():
            if value == target_value:
                return key
        return None


    def on_tick(self):
        # Pass time
        if (self.timer > 0):
            self.timer -= 1

        self.update_menu()
        root.after(1000, self.on_tick)
        

    def postpone(self):
        self.hide()
        self.timer = 60 * 5
        root.after(self.FIVE_MINUTES * 1000, self.check_and_show_popup)


    def check_and_show_popup(self):
            # Check if we are under working time range
            current_time = datetime.now().time()
            if self.start_time <= current_time <= self.end_time:
                self.load_task_list()
                self.show_popup()
            else:
                # Check again in a while
                root.after(self.popup_interval * 1000, self.check_and_show_popup)


    def show_popup(self):
        self.root.deiconify()
        self.root.attributes("-topmost", True)
        self.root.update_idletasks()

        # Move to the bottom right
        margin = 10
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        window_width = self.root.winfo_width()
        window_height = self.root.winfo_height()
        x = screen_width - window_width - 10
        y = screen_height - window_height - 10
        self.root.geometry(f"+{x-margin}+{y-margin}")

        for i in range(10):
            self.root.attributes("-alpha", i / 10)
            self.root.update()
            self.root.after(60)


    def load_task_list(self):
        try:
            with open("task_list.json") as file:
                self.task_list = json.load(file)

            # Load custom tasks from the json
            dropdown_values = list(self.task_list['CUSTOM_TASKS'].keys())

            # Load and append values from redmine
            if self.redmine_url != "''":
                self.load_redmine_issues()
                for key, value in self.redmine_issues.items():
                    dropdown_values.append(value)

            # Combine both and load them into the dropdown
            self.task_dropdown['values'] = dropdown_values
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load task list: {e}")


    def json_to_excel(self, json_file, excel_file):
        try:
            # Load the JSON data
            with open(json_file, 'r') as f:
                data = json.load(f)

            # Organize data by month
            monthly_data = {}
            for date_str, entries in data.items():
                date = datetime.strptime(date_str, '%d-%m-%Y')
                month_str = date.strftime('%B %Y')
                if month_str not in monthly_data:
                    monthly_data[month_str] = {}
                
                navision_data = entries.get("NAVISION_TIMESHEET", {})
                for phase, tasks in navision_data.items():
                    for task, hours in tasks.items():
                        if (phase, task) not in monthly_data[month_str]:
                            monthly_data[month_str][(phase, task)] = {}
                        monthly_data[month_str][(phase, task)][date.day] = hours

            # Create an Excel workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove the default sheet

            weekend_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for month, tasks in monthly_data.items():
                # Prepare data for the DataFrame
                rows = []
                for (phase, task), days in tasks.items():
                    row = {'PROJECT': 'SOUTHPAN', 'PHASE': phase, 'TASK': task, 'SUBPHASE': '', 'ON-CALL': '', 'CALL-IN': ''}
                    row.update({day: days.get(day, None) for day in range(1, 32)})
                    rows.append(row)
                
                # Create a DataFrame for the month
                df = pd.DataFrame(rows, columns=['PROJECT', 'PHASE', 'TASK', 'SUBPHASE', 'ON-CALL', 'CALL-IN'] + list(range(1, 32)))
                
                # Create a new sheet in the workbook
                ws = wb.create_sheet(title=month)
                
                # Write the DataFrame to the sheet
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        # Center align the header row
                        if r_idx == 1:
                            cell.alignment = Alignment(horizontal='center')
                        
                        # Apply number format to cells with numerical values (excluding the header row)
                        elif isinstance(value, (int, float)):
                            cell.number_format = '0.0'
                        
                        # Apply gray background to weekend columns
                        day_column = c_idx - 6  # Adjust for PROJECT, PHASE, TASK, SUBPHASE, ON-CALL, CALL-IN columns
                        if day_column > 0:
                            date_str = f"{day_column}-{month.split()[0]}-{month.split()[1]}"
                            try:
                                date = datetime.strptime(date_str, "%d-%B-%Y")
                                if date.weekday() >= 5:  # Saturday or Sunday
                                    cell.fill = weekend_fill
                                    if r_idx == 1:  # Apply the fill to the header
                                        cell.fill = weekend_fill
                            except ValueError:
                                continue
                        
                        # Apply the border style to all cells
                        cell.border = thin_border

                # Adjust column widths
                column_widths = {
                    'A': 10,  # PROJECT
                    'B': 10,  # PHASE
                    'C': 10,  # TASK
                    'D': 15,  # SUBPHASE
                    'E': 10,  # ON-CALL
                    'F': 10,  # CALL-IN
                }
                for col_idx, width in enumerate(column_widths.values(), 1):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = width

                # Adjust the widths of the day columns
                for col in range(7, 38):  # Columns for days 1 to 31
                    col_letter = get_column_letter(col)
                    ws.column_dimensions[col_letter].width = 3.5

            # Save the workbook to the specified file
            wb.save(excel_file)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to convert JSON to Excel: {e}")


    def hide(self):
        # Withdraw the Tkinter window
        self.root.withdraw()


    def show_tray_icon(self):
        # Load an image from a file
        image_path = "icon.png"
        image = Image.open(image_path)

        # Create the tray icon
        self.tray_icon = pystray.Icon("ptt_icon", image, "PTT - Productivity Tracking Tool v24.06.24")
        self.update_menu()
        self.tray_icon.run_detached()
    

    def update_menu(self):
        mins, secs = divmod(self.timer, 60)
        timeformat = '{:02d}:{:02d}'.format(mins, secs)
        
        # Define the menu items
        self.menu = (
            item(f'Time left: {timeformat}', self.do_nothing),
            item('Exit', self.on_closing),
        )
        self.tray_icon.menu = pystray.Menu(*self.menu)


    def restore_window(self, icon, item):
        self.tray_icon.stop()
        self.root.after(0, self.root.deiconify)


    def do_nothing(self):
        pass


    def on_closing(self):
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            self.tray_icon.stop()
        self.root.quit()


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
